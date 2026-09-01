import openpyxl
import sys
import json

sys.stdout.reconfigure(encoding='utf-8')

files = [
    ("Search Console CSV/advmdshahalam.me-Performance-on-Search-2026-09-01.xlsx", "FILE1_24hr"),
    ("Search Console CSV/advmdshahalam.me-Performance-on-Search-2026-09-01 (1).xlsx", "FILE2_7days"),
]

all_data = {}

for fpath, label in files:
    wb = openpyxl.load_workbook(fpath)
    all_data[label] = {}
    print(f"\n=== {label} ===")
    print(f"Sheets: {wb.sheetnames}")
    for sname in wb.sheetnames:
        ws = wb[sname]
        rows = list(ws.iter_rows(values_only=True))
        all_data[label][sname] = rows
        print(f"\n--- Sheet: {sname} (total rows: {len(rows)}) ---")
        for r in rows[:50]:
            print(r)

# Save as JSON for easy reading
import json
def default_serializer(obj):
    if hasattr(obj, 'isoformat'):
        return str(obj)
    return str(obj)

with open('excel_data.json', 'w', encoding='utf-8') as f:
    json.dump({k: {s: [list(r) for r in rows] for s, rows in v.items()} for k, v in all_data.items()}, f, ensure_ascii=False, default=default_serializer, indent=2)

print("\n\nSaved to excel_data.json")
